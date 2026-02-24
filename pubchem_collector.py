#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
pubchem_collector.py - PubChem Compound Excel Reporter

물질명을 입력받아 PubChem에서 연관된 모든 compound를 조회하고,
구조 이미지/SMILES/InChI/레퍼런스(논문·특허) 정보를 Excel로 저장합니다.

Usage:
    python pubchem_collector.py <compound_name>
    python pubchem_collector.py aspirin --max-cids 10
    python pubchem_collector.py              # 대화형 입력

출력: ./output/pubchem_{name}_{timestamp}.xlsx
"""

import sys
import io
import re
import json
import time
import argparse
import urllib.parse
import requests
from pathlib import Path
from datetime import datetime

# Windows 터미널 UTF-8 강제 적용
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ── 선택적 의존성 ─────────────────────────────────────────────────────────────
try:
    from rdkit import Chem
    from rdkit.Chem import AllChem
    from rdkit.Chem.Draw import rdMolDraw2D
    from rdkit.Chem.inchi import MolFromInchi
    RDKIT_AVAILABLE = True
except ImportError:
    RDKIT_AVAILABLE = False
    print("[!] RDKit 미설치 — structure_smiles / structure_inchi 이미지 생략됩니다.")

import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── 설정 상수 ─────────────────────────────────────────────────────────────────
PUBCHEM_BASE   = "https://pubchem.ncbi.nlm.nih.gov/rest/pug"
PUBCHEM_SDQ    = "https://pubchem.ncbi.nlm.nih.gov/sdq/sdqagent.cgi"

IMAGE_SIZE     = 250     # px
MAX_REFS       = 10      # compound당 최대 레퍼런스 수 (논문/특허 각각)
MAX_CIDS       = 20000    # 사실상 무제한
API_DELAY      = 0.25    # API 호출 간격 (초)

IMG_ROW_HEIGHT = 225     # pt  (300px ÷ 1.333 pt/px)
IMG_COL_WIDTH  = 40      # chr (300px ÷ 7 px/chr ≈ 43)
OUTPUT_DIR     = Path("output")

# 컬럼 인덱스 (1-based) — 총 19컬럼
COL_IMG_PUBCHEM = 11
COL_IMG_SMILES  = 12
COL_IMG_INCHI   = 13
COL_PAPER_DATE  = 16
COL_PAPER_DOI   = 17
COL_PATENT_DATE = 18
COL_PATENT_URL  = 19

# CAS 번호 패턴
_CAS_RE = re.compile(r"^\d{2,7}-\d{2}-\d$")

# 원소명 → 원소 기호 매핑 (SMARTS 서브구조 검색용)
ELEMENT_SYMBOL_MAP = {
    "hydrogen": "H",   "helium": "He",  "lithium": "Li",   "beryllium": "Be",
    "boron": "B",      "carbon": "C",   "nitrogen": "N",   "oxygen": "O",
    "fluorine": "F",   "neon": "Ne",    "sodium": "Na",    "magnesium": "Mg",
    "aluminum": "Al",  "aluminium": "Al","silicon": "Si",  "phosphorus": "P",
    "sulfur": "S",     "sulphur": "S",  "chlorine": "Cl",  "argon": "Ar",
    "potassium": "K",  "calcium": "Ca", "scandium": "Sc",  "titanium": "Ti",
    "vanadium": "V",   "chromium": "Cr","manganese": "Mn", "iron": "Fe",
    "cobalt": "Co",    "nickel": "Ni",  "copper": "Cu",    "zinc": "Zn",
    "gallium": "Ga",   "germanium": "Ge","arsenic": "As",  "selenium": "Se",
    "bromine": "Br",   "krypton": "Kr", "rubidium": "Rb",  "strontium": "Sr",
    "yttrium": "Y",    "zirconium": "Zr","niobium": "Nb",  "molybdenum": "Mo",
    "technetium": "Tc","ruthenium": "Ru","rhodium": "Rh",  "palladium": "Pd",
    "silver": "Ag",    "cadmium": "Cd", "indium": "In",    "tin": "Sn",
    "antimony": "Sb",  "tellurium": "Te","iodine": "I",    "xenon": "Xe",
    "cesium": "Cs",    "caesium": "Cs", "barium": "Ba",    "lanthanum": "La",
    "cerium": "Ce",    "praseodymium": "Pr","neodymium": "Nd","samarium": "Sm",
    "europium": "Eu",  "gadolinium": "Gd","terbium": "Tb", "dysprosium": "Dy",
    "holmium": "Ho",   "erbium": "Er",  "thulium": "Tm",   "ytterbium": "Yb",
    "lutetium": "Lu",  "hafnium": "Hf", "tantalum": "Ta",  "tungsten": "W",
    "rhenium": "Re",   "osmium": "Os",  "iridium": "Ir",   "platinum": "Pt",
    "gold": "Au",      "mercury": "Hg", "thallium": "Tl",  "lead": "Pb",
    "bismuth": "Bi",   "polonium": "Po","astatine": "At",  "radon": "Rn",
    "uranium": "U",    "thorium": "Th", "plutonium": "Pu",
}


# ═══════════════════════════════════════════════════════════════════════════════
# PubChem REST API
# ═══════════════════════════════════════════════════════════════════════════════

def _search_cids_smarts_async(smarts: str, max_results: int = 100000) -> list:
    """
    PubChem fastsubstructure SMARTS 검색으로 CID 목록을 반환합니다.
    비동기 ListKey 폴링 방식으로 대량 결과를 수집합니다.
    """
    print(f"  [*] SMARTS 서브구조 검색 시작: {smarts}")
    url = f"{PUBCHEM_BASE}/compound/fastsubstructure/smarts/cids/JSON"
    try:
        r = requests.post(url, data={"smarts": smarts}, timeout=60)
        data = r.json()

        # 즉시 결과 반환 (소규모)
        if "IdentifierList" in data:
            cids = data["IdentifierList"].get("CID", [])
            print(f"  [*] SMARTS 즉시 결과: {len(cids):,}개")
            return cids

        # ListKey 폴링 방식 (대규모)
        listkey = (data.get("Waiting", {}).get("ListKey")
                   or data.get("ListKey"))
        if not listkey:
            print(f"  [!] SMARTS 응답에서 ListKey 없음 (status={r.status_code})")
            return []

        print(f"  [*] ListKey={listkey}, 결과 수신 대기 중 ...")
        poll_url = (f"{PUBCHEM_BASE}/compound/listkey/{listkey}"
                    f"/cids/JSON?start=1&limit={max_results}")

        for attempt in range(60):   # 최대 120초 대기
            time.sleep(2)
            r2 = requests.get(poll_url, timeout=120)
            if r2.status_code == 200:
                data2 = r2.json()
                if "IdentifierList" in data2:
                    cids = data2["IdentifierList"].get("CID", [])
                    print(f"  [*] SMARTS 검색 완료: {len(cids):,}개")
                    return cids
                if "Waiting" in data2:
                    if attempt % 5 == 0:
                        print(f"  [*] 처리 중 ... ({attempt * 2}초 경과)")
                    continue
            elif r2.status_code == 404:
                print("  [!] ListKey 만료")
                break
            else:
                print(f"  [!] 폴링 응답 오류: {r2.status_code}")
                break

    except Exception as e:
        print(f"  [!] SMARTS 검색 오류: {e}")
    return []


def search_cids(name: str) -> list:
    """
    물질명으로 PubChem CID 목록을 다중 전략으로 반환합니다.

    전략 순서:
      1) 원소명 감지 → SMARTS [Symbol] 서브구조 검색 (가장 포괄적)
      2) name_type=word REST API (부분 일치)
      3) 정확 검색 fallback

    molybdenum처럼 원소명인 경우 SMARTS 경로를 통해 ~14,000개+를 수집합니다.
    """
    all_cids: set = set()

    # ── 전략 1: 원소명 감지 → SMARTS 서브구조 검색 ──────────────────────────
    element_symbol = ELEMENT_SYMBOL_MAP.get(name.lower().strip())
    if element_symbol:
        print(f"  [*] '{name}'은 원소명 → SMARTS 검색 사용 (기호: {element_symbol})")
        smarts_cids = _search_cids_smarts_async(f"[{element_symbol}]")
        if smarts_cids:
            all_cids.update(smarts_cids)
            print(f"  [*] 전략1(SMARTS): {len(smarts_cids):,}개 CID")

    # ── 전략 2: name_type=word (부분 일치) ──────────────────────────────────
    encoded = urllib.parse.quote(name)
    try:
        url = f"{PUBCHEM_BASE}/compound/name/{encoded}/cids/JSON?name_type=word"
        r = requests.get(url, timeout=30)
        if r.status_code == 200:
            cids = r.json().get("IdentifierList", {}).get("CID", [])
            if cids:
                new_count = len(set(cids) - all_cids)
                all_cids.update(cids)
                print(f"  [*] 전략2(word): {len(cids):,}개 CID (+신규 {new_count:,}개)")
    except Exception as e:
        print(f"  [!] word 검색 오류: {e}")

    if all_cids:
        result = sorted(all_cids)
        print(f"  [*] 최종 중복 제거 후: {len(result):,}개 CID")
        return result

    # ── 전략 3: 정확 검색 fallback ───────────────────────────────────────────
    try:
        url = f"{PUBCHEM_BASE}/compound/name/{encoded}/cids/JSON"
        r = requests.get(url, timeout=30)
        if r.status_code == 200:
            cids = r.json().get("IdentifierList", {}).get("CID", [])
            if cids:
                print(f"  [*] 전략3(정확검색): {len(cids):,}개 CID")
                return cids
    except Exception as e:
        print(f"  [!] 정확검색 오류: {e}")

    return []


def get_properties(cids: list) -> dict:
    """CID 목록의 compound 속성을 100개씩 배치 조회합니다."""
    result = {}
    for i in range(0, len(cids), 100):
        batch    = cids[i:i + 100]
        cids_str = ",".join(map(str, batch))
        props    = "CanonicalSMILES,IsomericSMILES,InChI,IUPACName,MolecularFormula,MolecularWeight"
        url      = f"{PUBCHEM_BASE}/compound/cid/{cids_str}/property/{props}/JSON"
        try:
            r = requests.get(url, timeout=30)
            r.raise_for_status()
            for p in r.json().get("PropertyTable", {}).get("Properties", []):
                result[p["CID"]] = p
        except Exception as e:
            print(f"  [!] get_properties batch {i // 100 + 1}: {e}")
        if i + 100 < len(cids):
            time.sleep(API_DELAY)
    return result


def get_synonyms(cids: list) -> dict:
    """CID별 동의어 목록을 50개씩 배치 조회합니다."""
    result = {}
    for i in range(0, len(cids), 50):
        batch    = cids[i:i + 50]
        cids_str = ",".join(map(str, batch))
        url      = f"{PUBCHEM_BASE}/compound/cid/{cids_str}/synonyms/JSON"
        try:
            r = requests.get(url, timeout=30)
            r.raise_for_status()
            for info in r.json().get("InformationList", {}).get("Information", []):
                result[info["CID"]] = info.get("Synonym", [])
        except Exception as e:
            print(f"  [!] get_synonyms batch {i // 50 + 1}: {e}")
        if i + 50 < len(cids):
            time.sleep(API_DELAY)
    return result


def get_pubchem_image(cid: int) -> bytes:
    """PubChem에서 2D 구조 PNG를 다운로드합니다."""
    url = f"{PUBCHEM_BASE}/compound/cid/{cid}/PNG?image_size={IMAGE_SIZE}x{IMAGE_SIZE}"
    try:
        r = requests.get(url, timeout=30)
        if r.status_code == 200:
            return r.content
    except Exception as e:
        print(f"  [!] get_pubchem_image CID {cid}: {e}")
    return None


# ── RDKit 이미지 ──────────────────────────────────────────────────────────────

def _mol_to_png(mol) -> bytes:
    if mol is None:
        return None
    try:
        AllChem.Compute2DCoords(mol)
        d = rdMolDraw2D.MolDraw2DCairo(IMAGE_SIZE, IMAGE_SIZE)
        d.DrawMolecule(mol)
        d.FinishDrawing()
        return d.GetDrawingText()
    except Exception as e:
        print(f"  [!] _mol_to_png: {e}")
        return None


def smiles_to_image(smiles: str) -> bytes:
    if not RDKIT_AVAILABLE or not smiles:
        return None
    try:
        return _mol_to_png(Chem.MolFromSmiles(smiles))
    except Exception:
        return None


def inchi_to_image(inchi: str) -> bytes:
    if not RDKIT_AVAILABLE or not inchi:
        return None
    try:
        return _mol_to_png(MolFromInchi(inchi))
    except Exception:
        return None


# ── SDQ API ───────────────────────────────────────────────────────────────────

def _sdq_query(collection: str, cid: int, limit: int = MAX_REFS) -> dict:
    """PubChem SDQ API로 레퍼런스를 최신순 조회합니다."""
    order_field = "articlepubdate" if collection == "pubmed" else "prioritydate"
    query = {
        "select": "*", "collection": collection,
        "where": {"ands": [{"cid": str(cid)}]},
        "order": [f"{order_field},desc"],
        "start": 1, "limit": limit,
    }
    params = {"infmt": "json", "outfmt": "json", "query": json.dumps(query)}
    try:
        r = requests.get(PUBCHEM_SDQ, params=params, timeout=60)
        if r.status_code == 200:
            out = r.json().get("SDQOutputSet", [{}])[0]
            return {"totalCount": out.get("totalCount", 0), "rows": out.get("rows", [])}
    except Exception as e:
        print(f"  [!] SDQ {collection}/CID {cid}: {e}")
    return {"totalCount": 0, "rows": []}


# ═══════════════════════════════════════════════════════════════════════════════
# 데이터 가공
# ═══════════════════════════════════════════════════════════════════════════════

def _smiles_key(prop: dict) -> str:
    return (prop.get("IsomericSMILES") or prop.get("CanonicalSMILES")
            or prop.get("SMILES") or "")


def extract_compound_name_and_synonyms(syns: list) -> tuple:
    """
    동의어 목록에서 compound_name(첫 번째 일반명)과
    compound_synonym(일반명 10개, 줄바꿈 구분)을 추출합니다.
    CAS 번호, InChI, 너무 긴 문자열은 제외합니다.
    """
    def _is_name(s: str) -> bool:
        s = s.strip()
        if not s:
            return False
        if _CAS_RE.match(s):          # CAS 번호
            return False
        if s.startswith("InChI="):    # InChI
            return False
        if s.startswith("[") and "]" in s and len(s) > 20:  # SMILES-like
            return False
        if len(s) > 100:              # 너무 긴 문자열
            return False
        return True

    names = [s for s in syns if _is_name(s)]
    compound_name    = names[0] if names else ""
    compound_synonym = "\n".join(names[1:11]) if len(names) > 1 else ""
    return compound_name, compound_synonym


def extract_cas(syns: list) -> str:
    """동의어 목록에서 첫 번째 CAS 번호를 추출합니다."""
    for s in syns:
        if _CAS_RE.match(str(s).strip()):
            return s.strip()
    return ""


def get_halogens(formula: str) -> str:
    """분자식에서 할로겐 원소(F, Cl, Br, I, At)를 추출합니다."""
    if not formula:
        return ""
    elements = set(re.findall(r"[A-Z][a-z]?", formula))
    found = [h for h in ["F", "Cl", "Br", "I", "At"] if h in elements]
    return ", ".join(found)


def process_paper_refs(rows: list) -> tuple:
    """
    pubmed SDQ rows -> (pubdate_str, doi_str)
    - pubdate_str : 최신순 날짜 줄바꿈 문자열
    - doi_str     : 최신순 DOI/PubMed URL 줄바꿈 문자열
    """
    dates, dois = [], []
    for ref in rows:
        date = ref.get("articlepubdate", "")
        doi  = ref.get("doi", "")
        pmid = ref.get("pmid", "")
        dates.append(str(date) if date else "")
        if doi:
            dois.append(str(doi))
        elif pmid:
            dois.append(f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/")
        else:
            dois.append("")
    return "\n".join(dates), "\n".join(dois)


def process_patent_refs(rows: list) -> tuple:
    """
    patent SDQ rows -> (pubdate_str, url_str)
    - pubdate_str : 최신순 날짜 줄바꿈 문자열
    - url_str     : Google Patents URL 줄바꿈 문자열
    """
    dates, urls = [], []
    for ref in rows:
        date    = ref.get("prioritydate", "")
        pub_num = ref.get("publicationnumber", "")
        url     = f"https://patents.google.com/patent/{pub_num.replace('-', '')}" if pub_num else ""
        dates.append(str(date) if date else "")
        urls.append(url)
    return "\n".join(dates), "\n".join(urls)


# ═══════════════════════════════════════════════════════════════════════════════
# Excel 스타일 (Arial Narrow 11pt)
# ═══════════════════════════════════════════════════════════════════════════════

_FONT_NAME = "Arial Narrow"
_FONT_SIZE = 11

HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(name=_FONT_NAME, size=_FONT_SIZE, bold=True, color="FFFFFF")
DATA_FONT  = Font(name=_FONT_NAME, size=_FONT_SIZE)
URL_FONT   = Font(name=_FONT_NAME, size=_FONT_SIZE, color="0563C1", underline="single")

WRAP_TOP   = Alignment(wrap_text=True, vertical="top",    horizontal="left")
CENTER_MID = Alignment(wrap_text=True, vertical="center", horizontal="center")
IMG_ALIGN  = Alignment(wrap_text=False, vertical="center",   horizontal="center")

_THIN       = Side(style="thin",   color="CCCCCC")
_MED        = Side(style="medium", color="888888")
THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# 컬럼 정의: (헤더 레이블, 너비)  — 19컬럼
COLUMNS = [
    ("CID",                          12),   # 1
    ("Compound Name",                25),   # 2
    ("Compound Synonym",             30),   # 3
    ("IUPAC Name",                   30),   # 4
    ("Molecular Formula",            16),   # 5
    ("Molecular Weight (g/mol)",     20),   # 6
    ("CAS",                          14),   # 7
    ("Halogen Elements",             14),   # 8
    ("SMILES",                       40),   # 9
    ("InChI",                        45),   # 10
    ("structure_pubchem", IMG_COL_WIDTH),   # 11
    ("structure_smiles",  IMG_COL_WIDTH),   # 12
    ("structure_inchi",   IMG_COL_WIDTH),   # 13
    ("paper_count",                  13),   # 14
    ("patent_count",                 13),   # 15
    ("paper_pubdate",                14),   # 16
    ("paper_doi",                    50),   # 17
    ("patent_pubdate",               14),   # 18
    ("patent_url",                   50),   # 19
]


# ═══════════════════════════════════════════════════════════════════════════════
# Excel 생성
# ═══════════════════════════════════════════════════════════════════════════════

def _add_image(ws, img_bytes: bytes, cell_ref: str):
    """PNG bytes를 Excel 셀에 부유 이미지로 삽입합니다."""
    if not img_bytes:
        return
    try:
        buf = io.BytesIO(img_bytes)
        xl  = XLImage(buf)
        xl.width  = IMAGE_SIZE
        xl.height = IMAGE_SIZE
        ws.add_image(xl, cell_ref)
    except Exception as e:
        print(f"  [!] 이미지 삽입 실패 {cell_ref}: {e}")


def _write_header(ws):
    """헤더 행(row=1)을 작성합니다."""
    for c_idx, (label, width) in enumerate(COLUMNS, start=1):
        cell           = ws.cell(row=1, column=c_idx, value=label)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = CENTER_MID
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(c_idx)].width = width
    ws.row_dimensions[1].height = 30


def _write_data_row(ws, row: int, cd: dict):
    """
    compound 1개를 단일 행으로 작성합니다.
    - 날짜/DOI/URL 셀: 줄바꿈 문자열(\\n) 사용, 개별 셀 분할 없음
    - 이미지 컬럼: 300x300 이미지 삽입
    """
    ws.row_dimensions[row].height = IMG_ROW_HEIGHT  # 225pt = 300px

    values = [
        cd["cid"],                           # 1  CID
        cd.get("compound_name", ""),         # 2  Compound Name
        cd.get("compound_synonym", ""),      # 3  Compound Synonym
        cd.get("iupac_name", ""),            # 4  IUPAC Name
        cd.get("formula", ""),              # 5  Molecular Formula
        cd.get("mw", ""),                   # 6  Molecular Weight
        cd.get("cas", ""),                  # 7  CAS
        cd.get("halogens", ""),             # 8  Halogen Elements
        cd.get("smiles", ""),               # 9  SMILES
        cd.get("inchi", ""),                # 10 InChI
        "",                                  # 11 structure_pubchem (image)
        "",                                  # 12 structure_smiles  (image)
        "",                                  # 13 structure_inchi   (image)
        cd.get("paper_count", 0),            # 14 paper_count
        cd.get("patent_count", 0),           # 15 patent_count
        cd.get("paper_pubdate_str", ""),     # 16 paper_pubdate
        cd.get("paper_doi_str", ""),         # 17 paper_doi
        cd.get("patent_pubdate_str", ""),    # 18 patent_pubdate
        cd.get("patent_url_str", ""),        # 19 patent_url
    ]

    for c_idx, val in enumerate(values, start=1):
        cell        = ws.cell(row=row, column=c_idx, value=val)
        cell.border = THIN_BORDER

        if c_idx in (COL_IMG_PUBCHEM, COL_IMG_SMILES, COL_IMG_INCHI):
            cell.font      = DATA_FONT
            cell.alignment = IMG_ALIGN
        elif c_idx in (COL_PAPER_DOI, COL_PATENT_URL):
            # URL/DOI 셀: 줄바꿈 표시, 첫 번째 URL에 셀 하이퍼링크
            cell.font      = URL_FONT
            cell.alignment = WRAP_TOP
            first = val.split("\n")[0].strip() if val else ""
            if first:
                # DOI면 doi.org URL로, URL이면 그대로
                if c_idx == COL_PAPER_DOI and not first.startswith("http"):
                    cell.hyperlink = f"https://doi.org/{first}"
                else:
                    cell.hyperlink = first
        else:
            cell.font      = DATA_FONT
            cell.alignment = WRAP_TOP

    # 이미지 삽입
    _add_image(ws, cd.get("img_pubchem"), f"{get_column_letter(COL_IMG_PUBCHEM)}{row}")
    _add_image(ws, cd.get("img_smiles"),  f"{get_column_letter(COL_IMG_SMILES)}{row}")
    _add_image(ws, cd.get("img_inchi"),   f"{get_column_letter(COL_IMG_INCHI)}{row}")


def build_excel(compound_data: list, output_path: Path) -> Path:
    """compound_data 목록을 Excel 파일로 저장합니다."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Compounds"

    _write_header(ws)
    ws.freeze_panes = "A2"

    for row_idx, cd in enumerate(compound_data, start=2):
        _write_data_row(ws, row_idx, cd)

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
# 메인 수집 흐름
# ═══════════════════════════════════════════════════════════════════════════════

def collect(name: str, skip_images: bool = False) -> list:
    """
    물질명에 대한 PubChem 데이터를 수집하여 compound_data 목록을 반환합니다.

    Args:
        name: 검색할 물질명
        skip_images: True이면 구조 이미지 다운로드/렌더링 생략 (대규모 수집 시 권장)
    """

    print(f"\n[1/6] '{name}' PubChem 검색 중 ...")
    cids = search_cids(name)
    if not cids:
        print(f"  [!] '{name}'에 해당하는 CID가 없습니다.")
        return []

    cids = cids[:MAX_CIDS]
    print(f"      {len(cids):,}개 CID 발견: {cids[:10]}{'...' if len(cids) > 10 else ''}")
    if skip_images:
        print("      [*] --skip-images 모드: 구조 이미지 생략")

    print(f"[2/6] Compound 속성 조회 중 ({len(cids):,}개) ...")
    props = get_properties(cids)

    print(f"[3/6] 동의어/CAS 조회 중 ({len(cids):,}개) ...")
    synonyms_map = get_synonyms(cids)

    compound_data = []
    total = len(cids)

    for i, cid in enumerate(cids, start=1):
        print(f"[4/6] CID {cid} ({i}/{total}) ...", end=" ")

        p      = props.get(cid, {})
        smiles = _smiles_key(p)
        inchi  = p.get("InChI", "")
        formula = p.get("MolecularFormula", "")
        syns   = synonyms_map.get(cid, [])

        compound_name, compound_synonym = extract_compound_name_and_synonyms(syns)
        cas      = extract_cas(syns)
        halogens = get_halogens(formula)

        # 구조 이미지 (--skip-images 시 생략)
        if skip_images:
            img_pubchem = img_smiles = img_inchi = None
        else:
            img_pubchem = get_pubchem_image(cid);  time.sleep(API_DELAY)
            img_smiles  = smiles_to_image(smiles)
            img_inchi   = inchi_to_image(inchi)

        # 논문/특허 레퍼런스
        paper_res  = _sdq_query("pubmed",  cid, limit=MAX_REFS);  time.sleep(API_DELAY)
        patent_res = _sdq_query("patent",  cid, limit=MAX_REFS);  time.sleep(API_DELAY)

        paper_pubdate_str,  paper_doi_str   = process_paper_refs(paper_res["rows"])
        patent_pubdate_str, patent_url_str  = process_patent_refs(patent_res["rows"])

        print(f"paper:{paper_res['totalCount']:,} / patent:{patent_res['totalCount']:,}")

        compound_data.append({
            # 기본 정보
            "cid":               cid,
            "compound_name":     compound_name,
            "compound_synonym":  compound_synonym,
            "iupac_name":        p.get("IUPACName", ""),
            "formula":           formula,
            "mw":                p.get("MolecularWeight", ""),
            "cas":               cas,
            "halogens":          halogens,
            "smiles":            smiles,
            "inchi":             inchi,
            # 구조 이미지
            "img_pubchem":       img_pubchem,
            "img_smiles":        img_smiles,
            "img_inchi":         img_inchi,
            # 레퍼런스 카운트
            "paper_count":       paper_res["totalCount"],
            "patent_count":      patent_res["totalCount"],
            # 레퍼런스 날짜/링크 (줄바꿈 문자열)
            "paper_pubdate_str":  paper_pubdate_str,
            "paper_doi_str":      paper_doi_str,
            "patent_pubdate_str": patent_pubdate_str,
            "patent_url_str":     patent_url_str,
        })

    return compound_data


def main():
    global MAX_CIDS, MAX_REFS

    parser = argparse.ArgumentParser(
        description="PubChem 물질 조회 -> Excel 저장 도구",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("name", nargs="?", help="검색할 물질명")
    parser.add_argument("--max-cids", type=int, default=None,
                        help="최대 CID 수 (기본: 무제한)")
    parser.add_argument("--max-refs", type=int, default=None,
                        help=f"compound당 최대 레퍼런스 수 (기본: {MAX_REFS})")
    parser.add_argument("--output", type=str, default=None,
                        help="저장 파일명 (미지정 시 자동 생성)")
    parser.add_argument("--skip-images", action="store_true",
                        help="구조 이미지 다운로드 생략 (대규모 수집 시 속도 향상)")
    args = parser.parse_args()

    if args.max_cids is not None:
        MAX_CIDS = args.max_cids
    if args.max_refs is not None:
        MAX_REFS = args.max_refs

    name = args.name or input("검색할 물질명을 입력하세요: ").strip()
    if not name:
        sys.exit("[!] 물질명이 입력되지 않았습니다.")

    compound_data = collect(name, skip_images=args.skip_images)
    if not compound_data:
        sys.exit("[!] 수집된 데이터가 없습니다.")

    if args.output:
        output_path = OUTPUT_DIR / args.output
    else:
        safe = "".join(c if c.isalnum() or c in " _-" else "_" for c in name)
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = OUTPUT_DIR / f"pubchem_{safe}_{ts}.xlsx"

    print(f"\n[5/6] Excel 파일 생성 중 -> {output_path} ...")
    build_excel(compound_data, output_path)

    print(f"\n[6/6] 완료!")
    print(f"      저장 위치 : {output_path.resolve()}")
    print(f"      수집 CID  : {len(compound_data)}개")
    print(f"      Excel 행수 : {len(compound_data) + 1}행 (헤더 포함)")
    if not RDKIT_AVAILABLE:
        print("      [주의] RDKit 미설치 -> structure_smiles / structure_inchi 이미지 비어있음")


if __name__ == "__main__":
    main()
